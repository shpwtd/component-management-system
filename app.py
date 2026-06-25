# -*- coding: utf-8 -*-
import os, json, datetime, io, re, time
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask import send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging
logging.basicConfig(filename="data/error.log", level=logging.ERROR, format="%(asctime)s %(levelname)s: %(message)s", encoding="utf-8")
from classifier import classify, detect_package, get_category_list, get_category_display, plan_layout
from classifier import get_category_tree

app = Flask(__name__)
app.secret_key = "component-mgr-secret-2024"
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
UPLOAD_FOLDER = os.path.join(DATA_DIR, "images")
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

def _load_json(name):
    path = os.path.join(DATA_DIR, name)
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8-sig") as f:
        return json.load(f)

def _save_json(name, data):
    """原子写入 JSON。

    步骤：写临时文件 -> fsync 落盘 -> os.replace 原子替换。
    任意一步失败都会保留损坏的临时文件（.broken）以便事后排查，
    不会覆盖原文件，杜绝"断电/进程崩溃导致 JSON 被写坏"的灾难。
    """
    path = os.path.join(DATA_DIR, name)
    tmp_path = f"{path}.{os.getpid()}.{int(time.time() * 1000)}.tmp"
    try:
        with open(tmp_path, "w", encoding="utf-8", newline="") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            f.flush()
            try:
                os.fsync(f.fileno())
            except OSError:
                # Windows 上某些 FS 不支持 fsync，忽略即可，不影响正确性
                pass
        os.replace(tmp_path, path)
    except Exception:
        # 写入或替换失败：保留坏文件以便排查，原文件不动
        if os.path.exists(tmp_path):
            try:
                os.replace(tmp_path, f"{tmp_path}.broken")
            except OSError:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
        logging.exception("原子写入 %s 失败", name)
        raise

def load_shelves():
    """读取货架列表，按 order 升序排序（order 决定首页显示和展开时的左右切分）。"""
    data = _load_json("shelves.json")
    data.sort(key=lambda s: (s.get("order", 0), s.get("id", 0)))
    return data

def save_shelves(data):
    _save_json("shelves.json", data)

def load_components():
    return _load_json("components.json")

def save_components(data):
    _save_json("components.json", data)

def load_logs():
    return _load_json("stock_logs.json")

def save_logs(data):
    _save_json("stock_logs.json", data)

def next_id(data):
    return max([x["id"] for x in data], default=0) + 1

def now_str():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def get_occupied(shelves, components):
    occ = {}
    for c in components:
        if c.get("shelf_id") and c.get("row") and c.get("col"):
            key = (c["shelf_id"], c["row"], c["col"])
            occ[key] = c
    return occ

def check_low_stock(components):
    return [c for c in components if c.get("min_stock", 0) > 0 and c["quantity"] <= c["min_stock"]]

def count_empty_slots(shelves, components):
    occ = get_occupied(shelves, components)
    total = 0
    for s in shelves:
        for r in s.get("rows", []):
            total += r["cols"]
    used = len(occ)
    return total - used, total

# ─── Home ─────────────────────────────────────────────────────
@app.route("/")
def index():
    shelves = load_shelves()
    components = load_components()
    logs = load_logs()
    occupied = get_occupied(shelves, components)
    low_stock = check_low_stock(components)
    used, total = count_empty_slots(shelves, components)
    top_10 = sorted(components, key=lambda c: sum(1 for l in logs if l.get("component_id") == c["id"] and l["type"] == "out"), reverse=True)[:10]
    shelf_usage = {}
    for s in shelves:
        total_slots = sum(rd["cols"] for rd in s.get("rows", []))
        used_slots = 0
        for rd in s.get("rows", []):
            for c in range(1, rd["cols"] + 1):
                if (s["id"], rd["row"], c) in occupied:
                    used_slots += 1
        pct = int(used_slots / total_slots * 100) if total_slots > 0 else 0
        shelf_usage[s["id"]] = {"used": used_slots, "total": total_slots, "pct": pct}
    return render_template("index.html", shelves=shelves, components=components, occupied=occupied,
                           low_stock=low_stock, used=used, total=total, top_10=top_10,
                           get_cat=get_category_display, shelf_usage=shelf_usage)

# ─── Shelves ─────────────────────────────────────────────────
@app.route("/shelves")
def shelf_list():
    shelves = load_shelves()
    components = load_components()
    occupied = get_occupied(shelves, components)
    return render_template("shelf_list.html", shelves=shelves, components=components, occupied=occupied)

@app.route("/shelves/add", methods=["GET", "POST"])
def shelf_add():
    shelves = load_shelves()
    default_order = max([s.get("order", 0) for s in shelves], default=0) + 1
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("请输入货架名称", "danger")
            return render_template("shelf_form.html", shelf=None, shelves=shelves, default_order=default_order)
        rows_data = []
        row_idx = 1
        while True:
            cols_str = request.form.get(f"cols_{row_idx}")
            if cols_str is None:
                break
            try:
                cols = int(cols_str)
                if cols > 0:
                    rows_data.append({"row": row_idx, "cols": cols})
            except ValueError:
                pass
            row_idx += 1
        if not rows_data:
            flash("请至少添加一行", "danger")
            return render_template("shelf_form.html", shelf=None, shelves=shelves, default_order=default_order)
        # 处理 order：默认 max+1，校验冲突
        existing_orders = [s.get("order") for s in shelves]
        try:
            order = int(request.form.get("order", str(default_order)).strip())
        except ValueError:
            order = default_order
        if order in existing_orders:
            flash(f"序号 {order} 已被占用，请换一个", "danger")
            return render_template("shelf_form.html", shelf=None, shelves=shelves, default_order=default_order)
        shelves.append({"id": next_id(shelves), "name": name, "rows": rows_data, "order": order})
        save_shelves(shelves)
        flash(f"货架「{name}」已创建", "success")
        return redirect(url_for("shelf_list"))
    return render_template("shelf_form.html", shelf=None, shelves=shelves, default_order=default_order)


@app.route("/uploads/<filename>")
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

@app.route("/shelves/<int:sid>/edit", methods=["GET", "POST"])
def shelf_edit(sid):
    shelves = load_shelves()
    shelf = next((s for s in shelves if s["id"] == sid), None)
    if not shelf:
        flash("货架不存在", "danger")
        return redirect(url_for("shelf_list"))
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("请输入货架名称", "danger")
            return render_template("shelf_form.html", shelf=shelf, shelves=shelves)
        rows_data = []
        row_idx = 1
        while True:
            cols_str = request.form.get(f"cols_{row_idx}")
            if cols_str is None:
                break
            try:
                cols = int(cols_str)
                if cols > 0:
                    rows_data.append({"row": row_idx, "cols": cols})
            except ValueError:
                pass
            row_idx += 1
        if not rows_data:
            flash("请至少添加一行", "danger")
            return render_template("shelf_form.html", shelf=shelf, shelves=shelves)
        # 处理 order：校验冲突（排除自己）
        existing_orders = [s.get("order") for s in shelves if s["id"] != sid]
        try:
            order = int(request.form.get("order", str(shelf.get("order", 1))).strip())
        except ValueError:
            order = shelf.get("order", 1)
        if order in existing_orders:
            flash(f"序号 {order} 已被占用，请换一个", "danger")
            return render_template("shelf_form.html", shelf=shelf, shelves=shelves)
        shelf["name"] = name
        shelf["rows"] = rows_data
        shelf["order"] = order
        save_shelves(shelves)
        flash(f"货架「{name}」已更新", "success")
        return redirect(url_for("shelf_list"))
    return render_template("shelf_form.html", shelf=shelf, shelves=shelves)

@app.route("/shelves/<int:sid>/delete", methods=["POST"])
def shelf_delete(sid):
    shelves = load_shelves()
    shelf = next((s for s in shelves if s["id"] == sid), None)
    if not shelf:
        flash("货架不存在", "danger")
        return redirect(url_for("shelf_list"))
    components = load_components()
    if any(c.get("shelf_id") == sid for c in components):
        flash(f"货架「{shelf['name']}」上还有元器件，请先移走再删除", "danger")
        return redirect(url_for("shelf_list"))
    shelves[:] = [s for s in shelves if s["id"] != sid]
    save_shelves(shelves)
    flash(f"货架「{shelf['name']}」已删除", "success")
    return redirect(url_for("shelf_list"))

@app.route("/shelves/<int:sid>/move", methods=["POST"])
def shelf_move(sid):
    """上下移：跟相邻柜子调换 order。load_shelves 已按 order 升序排好。"""
    shelves = load_shelves()
    idx = next((i for i, s in enumerate(shelves) if s["id"] == sid), None)
    if idx is None:
        flash("货架不存在", "danger")
        return redirect(url_for("shelf_list"))
    direction = request.form.get("dir", "")
    if direction == "up" and idx > 0:
        shelves[idx]["order"], shelves[idx - 1]["order"] = shelves[idx - 1]["order"], shelves[idx]["order"]
        save_shelves(shelves)
        flash(f"「{shelves[idx]['name']}」已上移", "success")
    elif direction == "down" and idx < len(shelves) - 1:
        shelves[idx]["order"], shelves[idx + 1]["order"] = shelves[idx + 1]["order"], shelves[idx]["order"]
        save_shelves(shelves)
        flash(f"「{shelves[idx]['name']}」已下移", "success")
    else:
        flash("无法移动（已在最上/最下）", "warning")
    return redirect(url_for("shelf_list"))

@app.route("/shelves/<int:sid>")
def shelf_detail(sid):
    shelves = load_shelves()
    shelf = next((s for s in shelves if s["id"] == sid), None)
    if not shelf:
        flash("货架不存在", "danger")
        return redirect(url_for("shelf_list"))
    components = [c for c in load_components() if c.get("shelf_id") == sid]
    return render_template("shelf_detail.html", shelf=shelf, components=components, get_cat=get_category_display)
# ─── Components ──────────────────────────────────────────────
@app.route("/components")
def component_list():
    components = load_components()
    components.sort(key=lambda c: (c.get("shelf_id") or 9999, c.get("row") or 9999, c.get("col") or 9999))
    shelves = load_shelves()
    shelf_map = {s["id"]: s["name"] for s in shelves}
    category = request.args.get("category", "")
    shelf_id = request.args.get("shelf_id", "")
    if category:
        components = [c for c in components if c.get("category") == category]
    if shelf_id:
        try:
            components = [c for c in components if c.get("shelf_id") == int(shelf_id)]
        except ValueError:
            pass
    return render_template("component_list.html", components=components, shelf_map=shelf_map,
                           get_cat=get_category_display, cats=get_category_list(), shelves=shelves,
                           cur_cat=category, cur_shelf=shelf_id)

@app.route("/components/add", methods=["GET", "POST"])
def component_add():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("请输入元器件名称", "danger")
            return render_template("component_form.html", comp=None, shelves=load_shelves(), cats=get_category_list(), tree=get_category_tree(), stock_logs=[])
        comp = {
            "id": 0,
            "name": name,
            "model": request.form.get("model", "").strip(),
            "package": request.form.get("package", "").strip(),
            "specs": request.form.get("specs", "").strip(),
            "quantity": int(request.form.get("quantity", 0)),
            "min_stock": int(request.form.get("min_stock", 0)),
            "category": request.form.get("category", "other"),
            "shelf_id": None, "row": None, "col": None,
            "created_at": now_str(), "updated_at": now_str()
        }
        shelf_id_str = request.form.get("shelf_id", "").strip()
        if shelf_id_str:
            try:
                sid = int(shelf_id_str)
                row = int(request.form.get("row", 0))
                col = int(request.form.get("col", 0))
                shelves = load_shelves()
                components = load_components()
                occupied = get_occupied(shelves, components)
                if (sid, row, col) in occupied:
                    flash("该货位已被占用", "danger")
                    return render_template("component_form.html", comp=comp, shelves=load_shelves(), cats=get_category_list(), tree=get_category_tree(), stock_logs=[])
                comp["shelf_id"] = sid
                comp["row"] = row
                comp["col"] = col
            except ValueError:
                pass
        components = load_components()
        comp["id"] = next_id(components)
        image_file = request.files.get("image")
        if image_file and image_file.filename:
            ext = image_file.filename.rsplit(".", 1)[1].lower() if "." in image_file.filename else ""
            if ext in ("jpg", "jpeg", "png", "gif", "webp"):
                fn = f"comp_{comp['id']}_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.{ext}"
                image_file.save(os.path.join(app.config["UPLOAD_FOLDER"], fn))
                comp["image"] = fn
        components.append(comp)
        save_components(components)
        flash(f"元器件「{name}」已添加", "success")
        return redirect(url_for("component_list"))
    return render_template("component_form.html", comp=None, shelves=load_shelves(), cats=get_category_list(), tree=get_category_tree(), stock_logs=[])

@app.route("/components/<int:cid>/edit", methods=["GET", "POST"])
def component_edit(cid):
    components = load_components()
    comp = next((c for c in components if c["id"] == cid), None)
    if not comp:
        flash("元器件不存在", "danger")
        return redirect(url_for("component_list"))
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("请输入元器件名称", "danger")
            return render_template("component_form.html", comp=comp, shelves=load_shelves(), cats=get_category_list(), tree=get_category_tree(), stock_logs=[])
        old_qty = comp["quantity"]
        comp["name"] = name
        comp["model"] = request.form.get("model", "").strip()
        comp["package"] = request.form.get("package", "").strip()
        comp["specs"] = request.form.get("specs", "").strip()
        comp["quantity"] = int(request.form.get("quantity", 0))
        comp["min_stock"] = int(request.form.get("min_stock", 0))
        comp["category"] = request.form.get("category", "other")
        comp["updated_at"] = now_str()
        shelf_id_str = request.form.get("shelf_id", "").strip()
        if shelf_id_str:
            try:
                sid = int(shelf_id_str)
                row = int(request.form.get("row", 0))
                col = int(request.form.get("col", 0))
                shelves = load_shelves()
                occupied = get_occupied(shelves, components)
                occ_key = (sid, row, col)
                if occ_key in occupied and occupied[occ_key]["id"] != cid:
                    flash("该货位已被占用", "danger")
                    return render_template("component_form.html", comp=comp, shelves=load_shelves(), cats=get_category_list(), tree=get_category_tree(), stock_logs=[])
                comp["shelf_id"] = sid
                comp["row"] = row
                comp["col"] = col
            except ValueError:
                pass
        # Log quantity change
        if comp["quantity"] != old_qty:
            logs = load_logs()
            diff = comp["quantity"] - old_qty
            logs.append({
                "id": next_id(logs),
                "component_id": cid,
                "type": "adjust",
                "quantity": abs(diff),
                "before": old_qty,
                "after": comp["quantity"],
                "note": "编辑时调整数量",
                "created_at": now_str()
            })
            save_logs(logs)
        image_file = request.files.get("image")
        if image_file and image_file.filename:
            ext = image_file.filename.rsplit(".", 1)[1].lower() if "." in image_file.filename else ""
            if ext in ("jpg", "jpeg", "png", "gif", "webp"):
                fn = f"comp_{cid}_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.{ext}"
                image_file.save(os.path.join(app.config["UPLOAD_FOLDER"], fn))
                comp["image"] = fn
        save_components(components)
        flash(f"元器件「{name}」已更新", "success")
        return redirect(url_for("component_list"))
    return render_template("component_form.html", comp=comp, shelves=load_shelves(), cats=get_category_list(), tree=get_category_tree(), stock_logs=[])

@app.route("/components/<int:cid>/delete", methods=["POST"])
def component_delete(cid):
    components = load_components()
    comp = next((c for c in components if c["id"] == cid), None)
    if not comp:
        flash("元器件不存在", "danger")
        return redirect(url_for("component_list"))
    components[:] = [c for c in components if c["id"] != cid]
    save_components(components)
    flash(f"元器件「{comp['name']}」已删除", "success")
    return redirect(url_for("component_list"))

@app.route("/components/<int:cid>")
def component_detail(cid):
    from classifier import get_category_list, get_category_tree
    components = load_components()
    comp = next((c for c in components if c["id"] == cid), None)
    if not comp:
        flash("元器件不存在", "danger")
        return redirect(url_for("component_list"))
    shelves = load_shelves()
    shelf_map = {s["id"]: s for s in shelves}
    logs = [l for l in load_logs() if l["component_id"] == cid]
    return render_template("component_form.html", comp=comp, shelves=shelves, cats=get_category_list(), stock_logs=logs, detail_view=True, tree=__import__("classifier").get_category_tree())
# ─── Stock In/Out ────────────────────────────────────────────
@app.route("/components/<int:cid>/stock/in", methods=["GET", "POST"])
def stock_in(cid):
    components = load_components()
    comp = next((c for c in components if c["id"] == cid), None)
    if not comp:
        flash("元器件不存在", "danger")
        return redirect(url_for("component_list"))
    if request.method == "POST":
        qty = int(request.form.get("quantity", 0))
        note = request.form.get("note", "").strip()
        if qty <= 0:
            flash("入库数量必须大于 0", "danger")
            return render_template("stock_in.html", comp=comp)
        logs = load_logs()
        before = comp["quantity"]
        comp["quantity"] += qty
        comp["updated_at"] = now_str()
        logs.append({
            "id": next_id(logs),
            "component_id": cid,
            "type": "in",
            "quantity": qty,
            "before": before,
            "after": comp["quantity"],
            "note": note,
            "created_at": now_str()
        })
        save_logs(logs)
        save_components(components)
        if comp.get("min_stock", 0) > 0 and comp["quantity"] <= comp["min_stock"]:
            flash(f"入库完成！注意：{comp['name']} 库存仍低于最低阈值", "warning")
        else:
            flash(f"入库 {qty} 个 {comp['name']} 完成", "success")
        return redirect(url_for("component_detail", cid=cid))
    return render_template("stock_in.html", comp=comp)

@app.route("/components/<int:cid>/stock/out", methods=["GET", "POST"])
def stock_out(cid):
    components = load_components()
    comp = next((c for c in components if c["id"] == cid), None)
    if not comp:
        flash("元器件不存在", "danger")
        return redirect(url_for("component_list"))
    if request.method == "POST":
        qty = int(request.form.get("quantity", 0))
        note = request.form.get("note", "").strip()
        if qty <= 0:
            flash("出库数量必须大于 0", "danger")
            return render_template("stock_out.html", comp=comp)
        if qty > comp["quantity"]:
            flash(f"库存不足！当前库存 {comp['quantity']} 个", "danger")
            return render_template("stock_out.html", comp=comp)
        logs = load_logs()
        before = comp["quantity"]
        comp["quantity"] -= qty
        comp["updated_at"] = now_str()
        logs.append({
            "id": next_id(logs),
            "component_id": cid,
            "type": "out",
            "quantity": qty,
            "before": before,
            "after": comp["quantity"],
            "note": note,
            "created_at": now_str()
        })
        save_logs(logs)
        save_components(components)
        if comp.get("min_stock", 0) > 0 and comp["quantity"] <= comp["min_stock"]:
            flash(f"出库完成！注意：{comp['name']} 库存已低于最低阈值 ({comp['quantity']}/{comp['min_stock']})", "warning")
        else:
            flash(f"出库 {qty} 个 {comp['name']} 完成", "success")
        return redirect(url_for("component_detail", cid=cid))
    return render_template("stock_out.html", comp=comp)

@app.route("/components/<int:cid>/history")
def stock_history(cid):
    components = load_components()
    comp = next((c for c in components if c["id"] == cid), None)
    if not comp:
        flash("元器件不存在", "danger")
        return redirect(url_for("component_list"))
    logs = [l for l in load_logs() if l["component_id"] == cid]
    logs.sort(key=lambda x: x["created_at"], reverse=True)
    return render_template("stock_history.html", comp=comp, logs=logs)
@app.route("/stock/logs")
def stock_logs_page():
    logs = load_logs()
    logs.sort(key=lambda x: x["created_at"], reverse=True)
    components = load_components()
    comp_map = {c["id"]: c["name"] for c in components}
    q = request.args.get("q", "").strip().lower()
    if q:
        logs = [l for l in logs if q in (comp_map.get(l.get("component_id", 0), "") or "").lower()]
    return render_template("stock_logs.html", logs=logs, comp_map=comp_map, q=q)
 
 
@app.route("/stock/logs/clear", methods=["POST"])
def stock_logs_clear():
    save_logs([])
    flash("所有流水记录已清空", "success")
    return redirect(url_for("stock_logs_page"))
 
 
# ─── Search ──────────────────────────────────────────────────
@app.route("/search")
def search():
    q = request.args.get("q", "").strip()
    results = []
    shelves = load_shelves()
    shelf_map = {s["id"]: s for s in shelves}
    if q:
        components = load_components()
        q_lower = q.lower()
        for c in components:
            if q_lower in c.get("name", "").lower() or q_lower in c.get("model", "").lower():
                results.append(c)
    return render_template("search.html", results=results, q=q, shelf_map=shelf_map, get_cat=get_category_display)

# ─── Smart Planning ──────────────────────────────────────────
@app.route("/plan", methods=["GET", "POST"])
def plan():
    shelves = load_shelves()
    components = load_components()
    if request.method == "POST":
        csv_text = request.form.get("csv_data", "").strip()
        if not csv_text:
            flash("请输入元器件数据", "danger")
            return render_template("plan.html", shelves=shelves, components=components, plan_result=None, get_cat=get_category_display)
        new_comps = []
        lines = csv_text.strip().split("\n")
        for line in lines:
            if not line.strip():
                continue
            parts = [p.strip() for p in line.split(",")]
            if len(parts) < 2:
                continue
            name = parts[0]
            model = parts[1] if len(parts) > 1 else ""
            pkg = parts[2] if len(parts) > 2 else ""
            specs = parts[3] if len(parts) > 3 else ""
            qty_str = parts[4] if len(parts) > 4 else "0"
            try:
                qty = int(qty_str)
            except ValueError:
                qty = 0
            cat_key = classify(name, "", specs, pkg)
            new_comps.append({"name": name, "model": model, "package": pkg, "specs": specs, "quantity": qty, "category": cat_key})
        result = plan_layout(new_comps, shelves)
        return render_template("plan.html", shelves=shelves, components=components, plan_result=result, get_cat=get_category_display)
    return render_template("plan.html", shelves=shelves, components=components, plan_result=None, get_cat=get_category_display)

@app.route("/plan/execute", methods=["POST"])
def plan_execute():
    data = request.get_json()
    if not data or "assignments" not in data:
        return jsonify({"success": False, "error": "无效的规划数据"}), 400
    components = load_components()
    for item in data["assignments"]:
        comp_info = item.get("component", {})
        comp = {
            "id": next_id(components),
            "name": comp_info.get("name", ""),
            "model": comp_info.get("model", ""),
            "package": comp_info.get("package", ""),
            "specs": comp_info.get("specs", ""),
            "quantity": int(comp_info.get("quantity", 0)),
            "min_stock": int(item.get("min_stock", 0)), "locked": int(item.get("locked", 0)),
            "category": comp_info.get("category", "other"),
            "shelf_id": item.get("shelf_id"),
            "row": item.get("row"),
            "col": item.get("col"),
            "created_at": now_str(),
            "updated_at": now_str()
        }
        components.append(comp)
    save_components(components)
    return jsonify({"success": True, "count": len(data["assignments"])})
# ─── Import / Export ──────────────────────────────────────────
@app.route("/import", methods=["GET", "POST"])
def import_page():
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            flash("请选择 Excel 文件", "danger")
            return redirect(url_for("import_page"))
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=3, max_col=9, values_only=True))
        except Exception as e:
            flash(f"文件解析失败：{e}", "danger")
            return redirect(url_for("import_page"))
        shelves = load_shelves()
        components = load_components()
        occupied = get_occupied(shelves, components)
        shelf_name_map = {s["name"]: s for s in shelves}
        direct_items = []
        planned_items = []
        conflict_items = []
        error_items = []
        for row_idx, row in enumerate(rows, start=3):
            if not any(row):
                continue
            name = str(row[0] or "").strip()
            specs = str(row[1] or "").strip()
            pkg = str(row[2] or "").strip()
            qty_str = str(row[3] or "0").strip()
            min_stock_str = str(row[4] or "0").strip()
            cat_name = str(row[5] or "").strip()
            shelf_name = str(row[6] or "").strip()
            row_num_str = str(row[7] or "").strip()
            col_str = str(row[8] or "").strip()
            try:
                qty = int(float(qty_str))
            except ValueError:
                qty = 0
            try:
                min_stock = int(float(min_stock_str))
            except ValueError:
                min_stock = 0
            if not name:
                error_items.append({"row": row_idx, "reason": "名称为空"})
                continue
            cat_list = get_category_list()
            cat_lookup = {cn: ck for ck, cn in cat_list}
            if cat_name in cat_lookup:
                cat_key = cat_lookup[cat_name]
            else:
                cat_key = classify(name, "", specs, pkg)
            if not shelf_name or not row_num_str or not col_str:
                assigned = False
                for s in shelves:
                    for rd in s.get("rows", []):
                        for c in range(1, rd["cols"]+1):
                            if (s["id"], rd["row"], c) not in occupied:
                                item = {"name": name, "model": "", "package": pkg, "specs": specs, "quantity": qty,
                                        "shelf_id": s["id"], "shelf_name": s["name"], "row": rd["row"], "col": c, "category": cat_key,
                                        "min_stock": min_stock, "locked": 0}
                                planned_items.append(item)
                                occupied[(s["id"], rd["row"], c)] = {"name": name}
                                assigned = True
                                break
                        if assigned: break
                    if assigned: break
                if not assigned:
                    error_items.append({"row": row_idx, "reason": "无可用的货位", "name": name})
                continue
            shelf = shelf_name_map.get(shelf_name)
            if not shelf:
                error_items.append({"row": row_idx, "reason": f"货架「{shelf_name}」不存在", "name": name})
                continue
            try:
                rn = int(row_num_str)
                cn = int(col_str)
            except ValueError:
                error_items.append({"row": row_idx, "reason": "行列号格式错误", "name": name})
                continue
            valid_row = any(rd["row"] == rn and cn <= rd["cols"] for rd in shelf.get("rows", []))
            if not valid_row:
                error_items.append({"row": row_idx, "reason": f"货架「{shelf_name}」不存在第{rn}行第{cn}列", "name": name})
                continue
            key = (shelf["id"], rn, cn)
            if key in occupied:
                conflict_items.append({"row": row_idx, "name": name, "shelf_name": shelf_name, "row": rn, "col": cn,
                                       "existing": occupied[key]["name"]})
                continue
            item = {
                "name": name, "model": "", "package": pkg, "specs": specs, "quantity": qty,
                "shelf_id": shelf["id"], "shelf_name": shelf_name, "row": rn, "col": cn, "category": cat_key,
                "min_stock": min_stock, "locked": 0
            }
            direct_items.append(item)
        return render_template("import_result.html", direct=direct_items, planned=planned_items, conflicts=conflict_items, errors=error_items,
                               total=len(rows))
    return render_template("import.html")

@app.route("/import/confirm", methods=["POST"])
def import_confirm():
    data = request.get_json()
    if not data or "items" not in data:
        return jsonify({"success": False, "error": "无导入数据"}), 400
    components = load_components()
    for item in data["items"]:
        comp = {
            "id": next_id(components),
            "name": item["name"],
            "model": "",
            "package": item.get("package", ""),
            "specs": item.get("specs", ""),
            "quantity": int(item.get("quantity", 0)),
            "min_stock": int(item.get("min_stock", 0)), "locked": int(item.get("locked", 0)),
            "category": item.get("category", "other"),
            "shelf_id": item.get("shelf_id"),
            "row": item.get("row"),
            "col": item.get("col"),
            "created_at": now_str(),
            "updated_at": now_str()
        }
        components.append(comp)
    save_components(components)
    return jsonify({"success": True, "count": len(data["items"])})

@app.route("/download_template")
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    headers = ["名称", "规格参数", "封装", "数量", "最低库存", "类别", "货架名称", "行号", "列号"]
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0984E3", end_color="0984E3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    # Row 2: Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 8
    # Row 1: Instruction
    cell = ws.cell(row=1, column=1, value="注意：第一行为说明行，不可删除。表头在第二行。从第三行开始填写数据。货架名称填实际名称（如A、B），不要加货架二字。如果不填货位（货架名称、行号、列号都留空），系统会自动分配货位。")
    cell.font = Font(bold=True, color="FF0000", size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    # Category dropdown (helper column K)
    from openpyxl.worksheet.datavalidation import DataValidation
    cats_path = os.path.join(DATA_DIR, "categories.json")
    if os.path.exists(cats_path):
        with open(cats_path, "r", encoding="utf-8-sig") as f:
            cats_data = json.load(f)
    else:
        cats_data = [{"key": "other", "name": "其他", "parent": None, "icon": "?"}]
    cat_names = [x["name"] for x in cats_data]
    for i_, name_ in enumerate(cat_names):
        ws.cell(row=3+i_, column=11, value=name_)
    ws.column_dimensions["K"].hidden = True
    last_row = 2 + len(cat_names)
    dv = DataValidation(type="list", formula1="$K$3:$K$" + str(last_row), allow_blank=True)
    dv.add("F3:F103")
    ws.add_data_validation(dv)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="元器件导入模板.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/export")
def export_excel():
    components = load_components()
    components.sort(key=lambda c: (c.get("shelf_id") or 9999, c.get("row") or 9999, c.get("col") or 9999))
    shelves = load_shelves()
    shelf_map = {s["id"]: s["name"] for s in shelves}
    wb = Workbook()
    ws = wb.active
    ws.title = "元器件列表"
    headers = ["名称", "规格参数", "封装", "数量", "最低库存", "类别", "货架", "行号", "列号"]
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="00B894", end_color="00B894", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    # Row 2: Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for row_idx, comp in enumerate(components, start=2):
        vals = [
            comp.get("name", ""),
            comp.get("specs", ""),
            comp.get("package", ""),
            comp.get("quantity", 0),
            comp.get("min_stock", 0),
            get_category_display(comp.get("category", "")),
            shelf_map.get(comp.get("shelf_id"), ""),
            comp.get("row", ""),
            comp.get("col", "")
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 8
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="元器件列表.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── AJAX helpers ───────────────────────────────────────────
@app.route("/api/shelf/<int:sid>/rows")
def api_shelf_rows(sid):
    shelves = load_shelves()
    shelf = next((s for s in shelves if s["id"] == sid), None)
    if not shelf:
        return jsonify([])
    return jsonify(shelf.get("rows", []))

@app.route("/api/shelf/<int:sid>/occupied")
def api_shelf_occupied(sid):
    components = [c for c in load_components() if c.get("shelf_id") == sid and c.get("row") and c.get("col")]
    result = {}
    for c in components:
        result[f"{c['row']}-{c['col']}"] = c["name"]
    return jsonify(result)


@app.route("/categories", methods=["GET", "POST"])
def categories_page():
    cats = __import__("classifier").load_categories()
    tree = __import__("classifier").get_category_tree()
    if __import__("flask").request.method == "POST":
        key = __import__("flask").request.form.get("key", "").strip()
        name = __import__("flask").request.form.get("name", "").strip()
        parent = __import__("flask").request.form.get("parent", "").strip() or None
        if not key or not name:
            __import__("flask").flash("请填写类别标识和名称", "danger")
            return __import__("flask").render_template("categories.html", cats=cats, tree=tree)
        key = __import__("re").sub(r"[^a-z0-9_]", "", key.lower())
        if not key:
            __import__("flask").flash("类别标识只能包含字母、数字和下划线", "danger")
            return __import__("flask").render_template("categories.html", cats=cats, tree=tree)
        if any(c["key"] == key for c in cats):
            __import__("flask").flash(f"类别标识「{key}」已存在", "danger")
            return __import__("flask").render_template("categories.html", cats=cats, tree=tree)
        cats.append({"key": key, "name": name, "parent": parent, "icon": "?"})
        __import__("json").dump(cats, open("data/categories.json","w",encoding="utf-8"), ensure_ascii=False, indent=2)
        __import__("flask").flash(f"类别「{name}」已添加", "success")
        return __import__("flask").redirect("/categories")
    return __import__("flask").render_template("categories.html", cats=cats, tree=tree)

@app.route("/categories/<cat_key>/delete", methods=["POST"])
def category_delete(cat_key):
    import json
    cats = __import__("classifier").load_categories()
    if cat_key in ("other",):
        __import__("flask").flash("不能删除「其他」类别", "danger")
        return __import__("flask").redirect("/categories")
    cat = next((c for c in cats if c["key"] == cat_key), None)
    if not cat:
        __import__("flask").flash("类别不存在", "danger")
        return __import__("flask").redirect("/categories")
    cats[:] = [c for c in cats if c["key"] != cat_key]
    comps = json.load(open("data/components.json","r",encoding="utf-8"))
    for comp in comps:
        if comp.get("category") == cat_key:
            comp["category"] = "other"
    json.dump(comps, open("data/components.json","w",encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(cats, open("data/categories.json","w",encoding="utf-8"), ensure_ascii=False, indent=2)
    __import__("flask").flash(f"类别「{cat['name']}」已删除", "success")
    return __import__("flask").redirect("/categories")



@app.route("/replan", methods=["GET"])
def replan_page():
    from classifier import suggest_group_key, get_category_display, get_category_list
    shelves = load_shelves()
    all_components = load_components()
    locked = [c for c in all_components if c.get("shelf_id") and c.get("locked")]
    comps = [c for c in all_components if not c.get("locked")]
    occ = get_occupied(shelves, all_components)
    if not comps:
        flash("没有可调整的元器件", "info")
        return render_template("replan.html", shelves=shelves, components=all_components, plan=None, locked_count=len(locked), get_cat=get_category_display, cats=get_category_list())
    groups = {}
    for c in comps:
        key = suggest_group_key(c.get("name",""),c.get("model",""),c.get("specs",""),c.get("package",""),c.get("category",""))
        if key not in groups: groups[key] = []
        groups[key].append(c)
    col_rows = {}
    for s in shelves:
        for rd in s.get("rows", []):
            cc = rd["cols"]
            if cc not in col_rows: col_rows[cc] = []
            col_rows[cc].append({"shelf_id": s["id"],"shelf_name": s["name"],"row": rd["row"],"cols": cc})
    plan = []
    moved = set()
    col_rows = {}
    for s in shelves:
        for rd in s.get("rows", []):
            cc = rd["cols"]
            if cc not in col_rows: col_rows[cc] = []
            col_rows[cc].append({"shelf_id": s["id"], "shelf_name": s["name"], "row": rd["row"], "cols": cc})

    for gk, grp in groups.items():
        if not grp: continue
        sm = grp[0]
        tcc = None
        if sm.get("shelf_id") and sm.get("row"):
            for s in shelves:
                m = [rd for rd in s.get("rows",[]) if rd["row"]==sm.get("row") and s["id"]==sm.get("shelf_id")]
                if m: tcc = m[0]["cols"]; break
        if not tcc or tcc not in col_rows:
            for cc in col_rows:
                tcc = cc
                break
        if not tcc or tcc not in col_rows: continue
        scored = [(ri, sum(1 for c in grp if c.get("shelf_id")==ri["shelf_id"] and c.get("row")==ri["row"]), sum(1 for col in range(1,ri["cols"]+1) if (ri["shelf_id"],ri["row"],col) not in occ and (ri["shelf_id"],ri["row"],col) not in moved)) for ri in col_rows[tcc]]
        scored.sort(key=lambda x: (-x[1], -x[2]))
        tgt = scored[0][0]
        for c in grp:
            if c["id"] in moved: continue
            if c.get("shelf_id")==tgt["shelf_id"] and c.get("row")==tgt["row"] and len(grp)>1:
                if any(o["id"]!=c["id"] and o.get("shelf_id")==tgt["shelf_id"] and o.get("row")==tgt["row"] for o in grp):
                    moved.add(c["id"]); continue
            tc = None
            for col in range(1, tgt["cols"]+1):
                if (tgt["shelf_id"],tgt["row"],col) not in occ:
                    if sum(1 for c2 in range(col+1,tgt["cols"]+1) if (tgt["shelf_id"],tgt["row"],c2) not in occ) > 0:
                        tc = col; break
            if not tc: continue
            plan.append({"component_id":c["id"],"component_name":c.get("name",""),"from_shelf":next((s["name"] for s in shelves if s["id"]==c.get("shelf_id")),""),"from_row":c.get("row"),"from_col":c.get("col"),"to_shelf":tgt["shelf_name"],"to_row":tgt["row"],"to_col":tc,"to_shelf_id":tgt["shelf_id"],"group":gk,"cols":tcc})
            moved.add(c["id"])
            occ[(tgt["shelf_id"],tgt["row"],tc)] = c
    plan.sort(key=lambda p: (p["group"],p["to_shelf"],p["to_row"],p["to_col"]))
    return render_template("replan.html", shelves=shelves, components=all_components, plan=plan, locked_count=len(locked), get_cat=get_category_display, cats=get_category_list())


@app.route("/replan/execute", methods=["POST"])
def replan_execute():
    data = request.get_json()
    if not data or "moves" not in data:
        return jsonify({"success": False, "error": "无效的调整数据"}), 400
    components = load_components()
    count = 0
    for m in data["moves"]:
        comp = next((c for c in components if c["id"] == m.get("component_id")), None)
        if not comp:
            continue
        comp["shelf_id"] = m.get("to_shelf_id")
        comp["row"] = m.get("to_row")
        comp["col"] = m.get("to_col")
        comp["updated_at"] = now_str()
        count += 1
    save_components(components)
    return jsonify({"success": True, "count": count})


@app.route("/components/<int:cid>/toggle_lock", methods=["POST"])
def component_toggle_lock(cid):
    comps = load_components()
    for c in comps:
        if c["id"] == cid:
            c["locked"] = 0 if c.get("locked") else 1
            c["updated_at"] = now_str()
            break
    save_components(comps)
    return jsonify({"success": True, "locked": next((c["locked"] for c in comps if c["id"] == cid), 0)})


@app.route("/components/batch_delete", methods=["POST"])
def component_batch_delete():
    data = request.get_json()
    if not data or "ids" not in data:
        return jsonify({"success": False, "error": "无效的请求数据"}), 400
    components = load_components()
    components[:] = [c for c in components if c["id"] not in data["ids"]]
    save_components(components)
    return jsonify({"success": True, "count": len(data["ids"])})


if __name__ == "__main__":
    # debug=False: 单机本地跑，不需要 Werkzeug 自动重载，反而少一个潜在风险面
    # 改了代码后手动重启即可（双击 start.vbs）
    app.run(host="127.0.0.1", port=5000, debug=False)
