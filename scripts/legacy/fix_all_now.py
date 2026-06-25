import sys, os
BASE = "."
path = os.path.join(BASE, "app.py")
with open(path, "r", encoding="utf-8") as f:
    app = f.read()

# Fix 1: component_add tree=
app = app.replace(
    'tree=get_category_tree(), stock_logs=',
    'tree=__import__("classifier").get_category_tree(), stock_logs='
)
print("Fix 1: component_add tree=")

# Fix 2: component_edit tree=
app = app.replace(
    'comp=comp, shelves=load_shelves(), cats=get_category_list(), stock_logs=[])',
    'comp=comp, shelves=load_shelves(), cats=get_category_list(), tree=__import__("classifier").get_category_tree(), stock_logs=[])'
)
print("Fix 2: component_edit tree=")

# Fix 3: shelf_add shelves=
app = app.replace(
    'return render_template("shelf_form.html", shelf=None)',
    'return render_template("shelf_form.html", shelf=None, shelves=load_shelves())'
)
print("Fix 3: shelf_add shelves=")

# Fix 4: DataValidation
marker = "ws.column_dimensions['I'].width = 8"
dv_code = """
    # Add category dropdown (helper column K)
    from openpyxl.worksheet.datavalidation import DataValidation
    cats = __import__("classifier").load_categories()
    cat_names = [x["name"] for x in cats]
    for i_, name_ in enumerate(cat_names):
        ws.cell(row=4+i_, column=11, value=name_)
    ws.column_dimensions["K"].hidden = True
    last_row = 3 + len(cat_names)
    dv = DataValidation(type="list", formula1="$K$4:$K$" + str(last_row), allow_blank=True)
    dv.add("F4:F100")
    ws.add_data_validation(dv)"""
if marker in app:
    app = app.replace(marker, marker + dv_code)
    print("Fix 4: DataValidation added")
else:
    print("Fix 4: WARNING - marker not found")

# Fix 5: Export sorting
old_export = 'components = load_components()\n    shelves = load_shelves()\n    shelf_map = {s["id"]: s["name"] for s in shelves}\n    wb = Workbook()'
new_export = 'components = load_components()\n    components.sort(key=lambda c: (c.get("shelf_id") or 9999, c.get("row") or 9999, c.get("col") or 9999))\n    shelves = load_shelves()\n    shelf_map = {s["id"]: s["name"] for s in shelves}\n    wb = Workbook()'
app = app.replace(old_export, new_export)
print("Fix 5: Export sorting")

with open(path, "w", encoding="utf-8") as f:
    f.write(app)

# Fix 6: Remove locked from component_form.html
tpl_path = os.path.join(BASE, "templates", "component_form.html")
with open(tpl_path, "r", encoding="utf-8") as f:
    tpl = f.read()

locked_idx = tpl.find('\u5e93\u4f4d\u9501\u5b9a')
if locked_idx >= 0:
    # Find the containing <div> block
    div_start = tpl.rfind('<div', 0, locked_idx)
    # Count nested divs
    depth = 0
    div_end = locked_idx
    for i in range(locked_idx, len(tpl)):
        if tpl[i:i+4] == '<div':
            depth += 1
        elif tpl[i:i+6] == '</div>':
            if depth == 0:
                div_end = i + 6
                break
            depth -= 1
    # If nested count didn't work, find next </div> after the <label>
    if div_end == locked_idx:
        div_end = tpl.find('</div>', locked_idx) + 6
        div_end = tpl.find('</div>', div_end) + 6
    
    tpl = tpl[:div_start] + tpl[div_end:]
    print("Fix 6: Locked section removed")
else:
    print("Fix 6: Locked section not found")

with open(tpl_path, "w", encoding="utf-8") as f:
    f.write(tpl)

# Verify
sys.path.insert(0, BASE)
from app import app as fa
print(f"\nOK! Routes: {len(list(fa.url_map.iter_rules()))}")
