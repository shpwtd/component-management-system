import sys, os, re
BASE = "."
path = os.path.join(BASE, "app.py")
with open(path, "r", encoding="utf-8") as f:
    app = f.read()

# ─── Fix 1: Export - remove model ───────────────────────────
old_export_vals = '''vals = [
            comp.get("name", ""),
            comp.get("model", ""),
            comp.get("package", ""),
            comp.get("specs", ""),
            comp.get("quantity", 0),
            comp.get("min_stock", 0),
            get_category_display(comp.get("category", "")),
            shelf_map.get(comp.get("shelf_id"), ""),
            comp.get("row", ""),
            comp.get("col", "")
        ]'''

new_export_vals = '''vals = [
            comp.get("name", ""),
            comp.get("specs", ""),
            comp.get("package", ""),
            comp.get("quantity", 0),
            comp.get("min_stock", 0),
            get_category_display(comp.get("category", "")),
            shelf_map.get(comp.get("shelf_id"), ""),
            comp.get("row", ""),
            comp.get("col", "")
        ]'''

if old_export_vals in app:
    app = app.replace(old_export_vals, new_export_vals)
    print("Fix 1: Export model removed")
else:
    print("Fix 1: WARNING - export vals pattern not found")

# ─── Fix 2: Template layout ────────────────────────────────
# 2a. Change headers from row=1 to row=2
old_header = 'for col_idx, h in enumerate(headers, 1):\n        cell = ws.cell(row=1, column=col_idx, value=h)\n        cell.font = header_font\n        cell.fill = header_fill\n        cell.alignment = Alignment(horizontal="center")\n        cell.border = thin_border'
new_header = '# Row 2: Headers\n    for col_idx, h in enumerate(headers, 1):\n        cell = ws.cell(row=2, column=col_idx, value=h)\n        cell.font = header_font\n        cell.fill = header_fill\n        cell.alignment = Alignment(horizontal="center")\n        cell.border = thin_border'

if old_header in app:
    app = app.replace(old_header, new_header)
    print("Fix 2a: Headers moved to row 2")
else:
    print("Fix 2a: WARNING - header pattern not found")

# 2b. Change example from row=2 to row=3
old_example = 'for col_idx, v in enumerate(example, 1):\n        cell = ws.cell(row=2, column=col_idx, value=v)\n        cell.border = thin_border'
new_example = 'for col_idx, v in enumerate(example, 1):\n        cell = ws.cell(row=3, column=col_idx, value=v)\n        cell.border = thin_border'

if old_example in app:
    app = app.replace(old_example, new_example)
    print("Fix 2b: Example moved to row 3")
else:
    print("Fix 2b: WARNING - example pattern not found")

# 2c. Replace instruction at row 3 with instruction at row 1
old_instr = 'ws.cell(row=3, column=1, value="填写说明：名称必填。填写货架名称+行号+列号则导入到指定货位，不填则自动分配空货位")'
new_instr = '''# Row 1: Instruction
    cell = ws.cell(row=1, column=1, value="注意：第一行为说明行，不可删除。从第三行开始填写数据。货架名称填实际名称（如A、B），不要加货架二字。")
    cell.font = Font(bold=True, color="FF0000", size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)'''

if old_instr in app:
    app = app.replace(old_instr, new_instr)
    print("Fix 2c: Instruction moved to row 1")
else:
    print("Fix 2c: WARNING - instruction pattern not found")

# 2d. Add DataValidation with helper column K
dv_marker = 'ws.column_dimensions["I"].width = 8'
dv_code = '''
    # Add category dropdown (helper column K)
    from openpyxl.worksheet.datavalidation import DataValidation
    cats = __import__("classifier").load_categories()
    cat_names = [x["name"] for x in cats]
    for i_, name_ in enumerate(cat_names):
        ws.cell(row=4+i_, column=11, value=name_)
    ws.column_dimensions["K"].hidden = True
    last_row = 3 + len(cat_names)
    dv = DataValidation(type="list", formula1="$K$4:$K$" + str(last_row), allow_blank=True)
    dv.add("F4:F100")
    ws.add_data_validation(dv)'''

if dv_marker in app:
    app = app.replace(dv_marker, dv_marker + dv_code)
    print("Fix 2d: DataValidation added (helper column K)")
else:
    print("Fix 2d: WARNING - column_dimensions I pattern not found")

with open(path, "w", encoding="utf-8") as f:
    f.write(app)

# Verify
sys.path.insert(0, BASE)
from app import app as fa
rules = [r.rule for r in fa.url_map.iter_rules()]
print(f"\nOK! Routes: {len(rules)}")
