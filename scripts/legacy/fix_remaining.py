with open("app.py","r",encoding="utf-8") as f:
    c = f.read()

# 1. Fix iter_rows: min_row=2 -> min_row=4, add max_col=9
c = c.replace(
    "rows = list(ws.iter_rows(min_row=2, values_only=True))",
    "rows = list(ws.iter_rows(min_row=4, max_col=9, values_only=True))"
)
print("1: iter_rows fixed (min_row=4, max_col=9)")

# 2. Fix example row to match new headers
old_example = 'example = ["\u8d34\u7247\u7535\u963b", "0805-10K", "0805", "10K\u03a9 \u00b11%", "100", "A\u8d27\u67b6", "1", "2"]'
new_example = 'example = ["\u8d34\u7247\u7535\u963b", "10K\u03a9 \u00b11%", "0805", "100", "50", "\u7535\u963b", "A", "1", "2"]'
c = c.replace(old_example, new_example)
print("2: example row fixed")

# 3. Add DataValidation 
dv_start = 'ws.column_dimensions["I"].width = 8'
dv_code = '''
    # Add category dropdown (hidden helper column K)
    from openpyxl.worksheet.datavalidation import DataValidation
    cats = __import__("classifier").load_categories()
    cat_names = [x["name"] for x in cats]
    for i, name in enumerate(cat_names):
        ws.cell(row=4+i, column=11, value=name)
    ws.column_dimensions["K"].hidden = True
    last_row = 3 + len(cat_names)
    dv = DataValidation(type="list", formula1="$K$4:$K$" + str(last_row), allow_blank=True)
    dv.add("F4:F100")
    ws.add_data_validation(dv)'''
c = c.replace(dv_start, dv_start + dv_code)
print("3: DataValidation added (helper column K)")

# 4. Fix replan route - remove duplicates, add clean version before if __name__
# Find and remove all replan_page routes, then add one clean one
import re

# Remove all existing @app.route("/replan") definitions
while '@app.route("/replan")' in c:
    start = c.index('@app.route("/replan")')
    # Find the end (next @app.route or def)
    rest = c[start+50:]
    end_match = re.search(r'\n@app\.route|\ndef ', rest)
    end = start + 50 + end_match.start() if end_match else len(c)
    c = c[:start] + c[end:]
    print(f"  Removed replan route from pos {start}")

# Add clean replan route before if __name__
if '@app.route("/replan", methods=[' not in c:
    replan_code = '''

@app.route("/replan", methods=["GET"])
def replan_page():
    shelves = load_shelves()
    all_components = load_components()
    locked = [c for c in all_components if c.get("shelf_id") and c.get("locked")]
    comps = [c for c in all_components if c.get("shelf_id") and not c.get("locked")]
    occ = get_occupied(shelves, all_components)
    if not comps:
        flash("没有可调整的元器件", "info")
        return render_template("replan.html", shelves=shelves, components=all_components, plan=None, locked_count=len(locked), get_cat=get_category_display, cats=get_category_list())
    groups = {}
    for c in comps:
        key = suggest_group_key(c.get("name",""),c.get("model",""),c.get("specs",""),c.get("package",""),c.get("category",""))
        if key not in groups: groups[key] = []
        groups[key].append(c)
    col_rows = {}
    for s in shelves:
        for rd in s.get("rows", []):
            cc = rd["cols"]
            if cc not in col_rows: col_rows[cc] = []
            col_rows[cc].append({"shelf_id": s["id"],"shelf_name": s["name"],"row": rd["row"],"cols": cc})
    plan = []
    moved = set()
    for gk, grp in sorted(groups.items(), key=lambda x: (-len(x[1]), x[0])):
        if not grp: continue
        sm = grp[0]
        tcc = None
        for s in shelves:
            m = [rd for rd in s.get("rows",[]) if rd["row"]==sm.get("row") and s["id"]==sm.get("shelf_id")]
            if m: tcc = m[0]["cols"]; break
        if not tcc or tcc not in col_rows: continue
        scored = [(ri, sum(1 for c in grp if c.get("shelf_id")==ri["shelf_id"] and c.get("row")==ri["row"]), sum(1 for col in range(1,ri["cols"]+1) if (ri["shelf_id"],ri["row"],col) not in occ and (ri["shelf_id"],ri["row"],col) not in moved)) for ri in col_rows[tcc]]
        scored.sort(key=lambda x: (-x[1], -x[2]))
        tgt = scored[0][0]
        for c in grp:
            if c["id"] in moved: continue
            if c.get("shelf_id")==tgt["shelf_id"] and c.get("row")==tgt["row"] and len(grp)>1:
                if any(o["id"]!=c["id"] and o.get("shelf_id")==tgt["shelf_id"] and o.get("row")==tgt["row"] for o in grp):
                    moved.add(c["id"]); continue
            tc = None
            for col in range(1, tgt["cols"]+1):
                if (tgt["shelf_id"],tgt["row"],col) not in occ:
                    if sum(1 for c2 in range(col+1,tgt["cols"]+1) if (tgt["shelf_id"],tgt["row"],c2) not in occ) > 0:
                        tc = col; break
            if not tc: continue
            plan.append({"component_id":c["id"],"component_name":c.get("name",""),"from_shelf":next((s["name"] for s in shelves if s["id"]==c.get("shelf_id")),""),"from_row":c.get("row"),"from_col":c.get("col"),"to_shelf":tgt["shelf_name"],"to_row":tgt["row"],"to_col":tc,"to_shelf_id":tgt["shelf_id"],"group":gk,"cols":tcc})
            moved.add(c["id"])
            occ[(tgt["shelf_id"],tgt["row"],tc)] = c
    plan.sort(key=lambda p: (p["group"],p["to_shelf"],p["to_row"],p["to_col"]))
    return render_template("replan.html", shelves=shelves, components=all_components, plan=plan, locked_count=len(locked), get_cat=get_category_display, cats=get_category_list())
'''
    c = c.replace("if __name__ == \"__main__\":", replan_code + "\n\nif __name__ == \"__main__\":")
    print("4: Replan route added")
else:
    print("4: Replan route already exists")

with open("app.py","w",encoding="utf-8") as f:
    f.write(c)

import sys; sys.path.insert(0,".")
from app import app as fa
print(f"\nRoutes: {len(list(fa.url_map.iter_rules()))}")
for r in sorted(fa.url_map.iter_rules()):
    if "replan" in r:
        print(f"  {r}")
