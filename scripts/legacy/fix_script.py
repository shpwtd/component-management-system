import json, os

app_path = "D:/codexspace/" + chr(20803) + chr(20214) + chr(20803) + chr(20215) + chr(31649) + chr(29702) + chr(31995) + "/app.py"

with open(app_path, "r", encoding="utf-8") as f:
    content = f.read()

# Fix download_template headers from row=1 back to row=2
old = "        cell = ws.cell(row=1, column=col_idx, value=h)\n        cell.font = header_font\n        cell.fill = header_fill\n        cell.alignment = Alignment(horizontal=\"center\")\n        cell.border = thin_border\n    ws.column_dimensions['A'].width"
new = "        cell = ws.cell(row=2, column=col_idx, value=h)\n        cell.font = header_font\n        cell.fill = header_fill\n        cell.alignment = Alignment(horizontal=\"center\")\n        cell.border = thin_border\n    ws.column_dimensions['A'].width"
content = content.replace(old, new)

# Add DataValidation + helper column K before "    bio = io.BytesIO()"
insert_before = "    bio = io.BytesIO()"
insert_code = """    # Category dropdown (helper column K)
    from openpyxl.worksheet.datavalidation import DataValidation
    cats_path = os.path.join(DATA_DIR, "categories.json")
    if os.path.exists(cats_path):
        with open(cats_path, "r", encoding="utf-8-sig") as f:
            cats_data = json.load(f)
    else:
        cats_data = [{"key": "other", "name": "其他", "parent": None, "icon": "?"}]
    cat_names = [x["name"] for x in cats_data]
    for i_, name_ in enumerate(cat_names):
        ws.cell(row=4+i_, column=11, value=name_)
    ws.column_dimensions["K"].hidden = True
    last_row = 3 + len(cat_names)
    dv = DataValidation(type="list", formula1="$K$4:$K$" + str(last_row), allow_blank=True)
    dv.add("F4:F104")
    ws.add_data_validation(dv)

"""
content = content.replace(insert_before, insert_code + insert_before)

with open(app_path, "w", encoding="utf-8") as f:
    f.write(content)

print("app.py fixed OK")
